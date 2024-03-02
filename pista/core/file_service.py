import ftplib
import json
import os
import socket
import threading
import time
from zipfile import BadZipFile

import openpyxl
import xml.etree.ElementTree as ET

import pandas as pd
from jsonschema.exceptions import ValidationError
from jsonschema.validators import validate
from openpyxl.worksheet.worksheet import Worksheet
from datetime import datetime, timedelta
from cryptography.fernet import Fernet

from core.common_service import Commons
from core.log_service import printit

from root import OUTPUT_DIR, SCREENSHOT_DIR, RUNTIME_DIR, TEST_DIR, DATAPREP_FILE


class ExcelUtil:
    @staticmethod
    def read_xlsheet(filepath: str, sheet) -> Worksheet:
        """ Get worksheet obj from excel file """
        try:
            workbook = openpyxl.load_workbook(filepath)
            if sheet is not None and sheet in workbook.sheetnames:
                print('Reading worksheet: ' + sheet)
            else:
                sheet = workbook.sheetnames[0]
                print('Worksheet not provided, considering: ' + sheet)

            worksheet = workbook[sheet]
        except FileNotFoundError as e:
            assert False, 'File not found: ' + filepath
        return worksheet

    @classmethod
    def read_all_xlrows_for_column(cls, filepath: str, col_header: str, row_header_to_avoid=None) -> str:
        """Returns str with values as csv from all the rows
        """
        thread_id = threading.current_thread().native_id
        col_values_str = ''

        col_hdr_num_dict, row_hdr_num_dict, max_row, max_col = cls._get_all_xlheaders_num_dict(filepath)

        workbook = None
        try:
            workbook = openpyxl.load_workbook(filepath, read_only=False)
            sheet = workbook.sheetnames[0]

            if row_header_to_avoid is not None and str(row_header_to_avoid) in row_hdr_num_dict.keys():
                row_to_avoid = row_hdr_num_dict[str(row_header_to_avoid)]
            else:
                row_to_avoid = None

            for i in range(2, max_row + 1):
                if row_to_avoid is not None and f"{row_to_avoid}" == f"{i}":
                    continue
                    
                col = col_hdr_num_dict[col_header]
                curr_cell_val = workbook[sheet].cell(row=i, column=col).value
                printit(f"xl ({i},{col}) {curr_cell_val}", isToPrint=False)
                
                if curr_cell_val is not None:
                    if col_values_str == '':
                        col_values_str = curr_cell_val
                    else:
                        col_values_str = col_values_str + ',' + curr_cell_val
        finally:
            if workbook is not None:
                workbook.close()

        # for column in workbook[sheet].iter_cols():
        #     column_name = column[0].value
        #     if column_name == col_header:
        #         for i, cell in enumerate(column):
        #             if i == 0:
        #                 continue
        #             if cell.value is not None and str(cell.value).strip() != '':
        #                 col_values_str = col_values_str + ',' + cell.value

        # print(Commons.build_date_forfilename(), f"{thread_id} reading xl (All,{col_header}): {col_values_str}")
        printit(f"... Thread {thread_id} reading excel data (ALL,{col_header}): {col_values_str}")
        return col_values_str

    @classmethod
    def _get_all_xlheaders_num_dict(cls, filepath: str, sheet:str=None):
        """Gets 2 dicts: column header name with seq num, row header name with seq num
        """
        col_hdr_num_dict = dict()
        row_hdr_num_dict = dict()

        workbook = None
        max_row = None
        max_col = None

        for _ in range(5):
            try:
                workbook = openpyxl.load_workbook(filepath, read_only=True)
                break
            except BadZipFile as e:
                time.sleep(2.0)

        try:
            if workbook is not None:
                sheet = workbook.sheetnames[0] if sheet is None else sheet

                max_col = workbook[sheet].max_column
                max_row = workbook[sheet].max_row

                for i in range(2, max_col + 1):
                    col_header = workbook[sheet].cell(row=1, column=i).value
                    col_hdr_num_dict[col_header] = i

                for i in range(2, max_row + 1):
                    row_header = workbook[sheet].cell(row=i, column=1).value
                    row_hdr_num_dict[str(row_header)] = i
        finally:
            if workbook is not None:
                workbook.close()

        return col_hdr_num_dict, row_hdr_num_dict, max_row, max_col

    @classmethod
    def _get_all_xlheaders_numrange_dict(cls, filepath: str, sheet:str=None):
        """Gets 2 dicts: column header name with seq num, row header name with seq num
        """
        col_hdr_num_dict = dict()
        row_hdr_num_dict = dict()

        workbook = None
        max_row = None

        for _ in range(5):
            try:
                workbook = openpyxl.load_workbook(filepath, read_only=True)
                break
            except BadZipFile as e:
                time.sleep(2.0)

        try:
            if workbook is not None:
                sheet = workbook.sheetnames[0] if sheet is None else sheet

                max_col = workbook[sheet].max_column
                max_row = workbook[sheet].max_row

                i = 2
                while i <= max_col:
                    start = end = i
                    col_header = workbook[sheet].cell(row=1, column=i).value
                    for j in range(start+1, max_col + 1):
                        extra_col_header = workbook[sheet].cell(row=1, column=j).value
                        if col_header == extra_col_header:
                            end = j
                        else:
                            end = start if end is None else end
                            break
                    i = end + 1
                    col_hdr_num_dict[col_header] = (start, end)

                i = 2
                while i <= max_row:
                    start = end = i
                    row_header = workbook[sheet].cell(row=i, column=1).value
                    for j in range(start + 1, max_row + 1):
                        extra_row_header = workbook[sheet].cell(row=j, column=1).value
                        if row_header == extra_row_header:
                            end = j
                        else:
                            end = start if end is None else end
                            break
                    i = end + 1
                    row_hdr_num_dict[row_header] = (start, end)
        finally:
            if workbook is not None:
                workbook.close()

        return col_hdr_num_dict, row_hdr_num_dict, max_row

    @staticmethod
    def append_to_xlcell(filepath: str, row_header, col_header: str, cell_value):
        """Appends str to existing cell value
        """
        thread_id = threading.current_thread().native_id

        col_hdr_num_dict, row_hdr_num_dict, max_row, max_col = ExcelUtil._get_all_xlheaders_num_dict(filepath)

        workbook = None

        for _ in range(5):
            try:
                workbook = openpyxl.load_workbook(filepath)
                break
            except BadZipFile as e:
                time.sleep(2.0)

        try:
            # workbook = openpyxl.load_workbook(filepath)
            sheet = workbook.active

            col_hdr_num = col_hdr_num_dict[col_header]
            if str(row_header) in row_hdr_num_dict.keys():
                row_hdr_num = row_hdr_num_dict[str(row_header)]
            else:
                new_row_num = max_row + 1
                # sheet.cell(row=new_row_num, column=1).value = str(row_header)
                sheet.append([str(row_header), '', '', ''])
                row_hdr_num = new_row_num

            curr_cell_val = sheet.cell(row=row_hdr_num, column=col_hdr_num).value
            if curr_cell_val is None or str(curr_cell_val).strip() == '':
                new_cell_val = str(cell_value)
            else:
                new_cell_val = str(curr_cell_val).strip() + ',' + str(cell_value)

            # print(Commons.build_date_forfilename(), f"{thread_id} updating xl ({row_header},{col_header}): {cell_value}")
            printit(f"... Thread {thread_id} updating excel data ({row_header},{col_header}): {cell_value}")
            sheet.cell(row=row_hdr_num, column=col_hdr_num).value = new_cell_val
        finally:
            if workbook is not None:
                workbook.save(filepath)
                workbook.close()

    @staticmethod
    def clear_xlcells(filepath: str, row_header, col_header: str = None):
        """"""
        thread_id = threading.current_thread().native_id

        col_hdr_num_dict, row_hdr_num_dict, max_row, max_col = ExcelUtil._get_all_xlheaders_num_dict(filepath)

        workbook = None

        for _ in range(5):
            try:
                workbook = openpyxl.load_workbook(filepath)
                break
            except BadZipFile as e:
                time.sleep(2.0)

        try:
            # workbook = openpyxl.load_workbook(filepath)
            sheet = workbook.active

            row_hdr_num = row_hdr_num_dict[str(row_header)] if str(row_header) in row_hdr_num_dict.keys() else None

            if row_hdr_num is not None:
                if col_header is not None:
                    '''Clear only cell for 1 row & 1 column'''
                    # print(Commons.build_date_forfilename(), f"{thread_id} clearing xl ({row_header},{col_header})")
                    printit(f"... Thread {thread_id} clearing excel data ({row_header},{col_header})")

                    col_hdr_num = col_hdr_num_dict[col_header] if str(col_header) in col_hdr_num_dict.keys() else None
                    if col_hdr_num is not None:
                        printit(f"xl ({row_hdr_num},{col_hdr_num})", isToPrint=False)
                        sheet.cell(row=row_hdr_num, column=col_hdr_num).value = ''
                else:
                    '''Clear all cells for 1 row'''
                    # print(Commons.build_date_forfilename(), f"{thread_id} clearing xl ({row_header},)")
                    printit(f"... Thread {thread_id} clearing excel data ({row_header},ALL)")
                    for i in range(2, max_col + 1):
                        printit(f"xl ({row_hdr_num},{i})", isToPrint=False)
                        sheet.cell(row=row_hdr_num, column=i).value = ''
        finally:
            if workbook is not None:
                workbook.save(filepath)
                workbook.close()

    @staticmethod
    def get_dataprep_input(file_path=DATAPREP_FILE, sheet='Sheet1'):
        import pandas as pd
        import numpy as np
        # Replace 'Sheet1' with the name of your sheet
        df = pd.read_excel(file_path, sheet_name=sheet)
        # Convert all NaN values to None
        df = df.replace(np.nan, None)
        # Group by 'Name' column and return a dictionary
        input_data = {name: [tuple(x)[2:] for x in group.itertuples()] for name, group in df.groupby('TEST_ID')}

        for key in input_data:
            new_list = []
            for item in input_data[key]:
                repeat = int(item[0]) if item[0] is not None else 1
                new_list.extend([item] * repeat)
            input_data[key] = new_list

        return input_data
    
# filepath = 'D:/Practice/AutomationService/resources/data/data_variable.xlsx'
# ws = Excel.read_sheet(filepath, 'DO')
# print(ws)


class ExcelDFUtil:

    @classmethod
    def _get_excel_as_df(cls, file_path:str, sheet_name:str):
        # Load the spreadsheet
        xls = pd.ExcelFile(file_path)
        # Load a sheet into a DataFrame by its name
        df = xls.parse(sheet_name)
        df = df.astype(str)
        df = df.replace({'nan': ''})

        return df

    @classmethod
    def _get_df_attr(cls, df):
        max_row, max_col = df.shape

        row_header_dict = df[df.columns[0]].to_dict()
        final_row_header_dict = {v: k for k, v in row_header_dict.items()}

        final_col_header_dict = {item: index for index, item in enumerate(df.columns.values)}

        return final_row_header_dict, final_col_header_dict, max_row, max_col

    @classmethod
    def read_from_excel(cls, file_path:str, sheet_name:str, row:int, column:int):
        df = cls._get_excel_as_df(file_path, sheet_name)

        # Read data from a particular cell
        data = df.iat[row, column]

        # data = int(float(data)) if Commons.check_number_type(data) == 'float' else data
        data = int(float(data)) if str(data).endswith('.0') else data
        
        return str(data)

    @classmethod
    def read_from_excel_with_header(cls, file_path, sheet_name, row_header, column_header):
        df = cls._get_excel_as_df(file_path, sheet_name)

        row_header_dict, col_header_dict, max_row, max_col = cls._get_df_attr(df)
        row = row_header_dict[row_header]

        # Read data from a particular cell
        data = df.loc[row, column_header]

        # data = int(float(data)) if Commons.check_number_type(data) == 'float' else data
        data = int(float(data)) if str(data).endswith('.0') else data

        return str(data)

    @classmethod
    def read_all_xlrows_for_column(cls, filepath: str, col_header: str) -> str:
        """Returns str with values as csv from all the rows
        """
        thread_id = threading.current_thread().native_id
        col_values_str_as_csv = ''

        df = cls._get_excel_as_df(filepath, 'Sheet1')
        row_hdr_num_dict, col_header_dict, max_row, max_col = cls._get_df_attr(df)

        try:
            for i in range(0, max_row):
                curr_cell_val = df.loc[i, col_header]
                printit(f"xl ({i},{col_header}) {curr_cell_val}", isToPrint=False)

                if curr_cell_val and curr_cell_val.lower() != 'nan' and curr_cell_val.strip() != '':
                    if col_values_str_as_csv == '':
                        col_values_str_as_csv = curr_cell_val
                    else:
                        col_values_str_as_csv += ',' + curr_cell_val
        finally:
            pass

        print(Commons.build_date_forfilename(), f"{thread_id} reading xl (All,{col_header}): {col_values_str_as_csv}")
        return col_values_str_as_csv

    @classmethod
    def write_to_excel(cls, file_path: str, sheet_name: str, row: int, column: int, data):
        # data = np.nan if data is None or data == '' else str(data)
        data = int(float(data)) if str(data).endswith('.0') else data

        df = cls._get_excel_as_df(file_path, sheet_name)

        # Write data to a particular cell
        df.iat[row, column] = data
        # Write DataFrame back to Excel file
        df.to_excel(file_path, sheet_name=sheet_name, index=False)

    @classmethod
    def write_to_excel_with_header(cls, file_path: str, sheet_name: str, row_header: str, column_header: str, data):
        # data = np.nan if data is None or data == '' else str(data)
        data = int(float(data)) if str(data).endswith('.0') else data

        df = cls._get_excel_as_df(file_path, sheet_name)

        row_header_dict, col_header_dict, max_row, max_col = cls._get_df_attr(df)
        row = row_header_dict[row_header]

        # Write data to a particular cell
        df.loc[row, column_header] = data
        # Write DataFrame back to Excel file
        df.to_excel(file_path, sheet_name=sheet_name, index=False)

    @classmethod
    def append_to_xlcell(cls, filepath: str, row_header, col_header: str, cell_value):
        """Appends str to existing cell value
        """
        try:
            # TODO write a new line for THREAD_ID if not found
            df = cls._get_excel_as_df(filepath, 'Sheet1')

            row_header_dict, col_header_dict, max_row, max_col = cls._get_df_attr(df)

            # Write new row_header to excel
            if str(row_header) not in row_header_dict.keys():
                df.loc[len(df.index)] = [row_header, '', '', '', '', '']
                df.to_excel(filepath, 'Sheet1', index=False)

            curr_val = cls.read_from_excel_with_header(filepath, 'Sheet1', row_header, col_header)
            if curr_val and len(curr_val) > 0 and curr_val.strip != '':
                new_val = curr_val + ',' + cell_value
            else:
                new_val = cell_value
            cls.write_to_excel_with_header(filepath, 'Sheet1', row_header, col_header, new_val)
        finally:
            pass

    @classmethod
    def clear_xlcells(cls, filepath: str, row_header, col_header: str = None):
        """"""
        thread_id = threading.current_thread().native_id

        try:
            df = cls._get_excel_as_df(filepath, 'Sheet1')

            row_header_dict, col_header_dict, max_row, max_col = cls._get_df_attr(df)
            row = row_header_dict[row_header]

            if col_header:
                '''Clear only cell for 1 row & 1 column'''
                print(Commons.build_date_forfilename(), f"{thread_id} clearing xl ({row_header},{col_header})")

                printit(f"xl ({row_header},{col_header})", isToPrint=False)
                cls.write_to_excel_with_header(filepath, 'Sheet1', row_header, col_header, '')
            else:
                '''Clear all cells for 1 row'''
                print(Commons.build_date_forfilename(), f"{thread_id} clearing xl ({row_header},)")

                for c in range(1, max_col):
                    printit(f"xl ({row},{c})", isToPrint=False)
                    cls.write_to_excel(filepath, 'Sheet1', row, c, '')
        finally:
            pass


class XMLUtil:
    @staticmethod
    def format_xml(xml_str: str) -> str:
        """ To get formatted xml from an xml str """
        formatd_xml = None
        try:
            xmlobj = ET.XML(xml_str)
            ET.indent(xmlobj)
            formatd_xml = ET.tostring(xmlobj, encoding='unicode')
        except Exception as e:
            print('Exception during xml formatting: ' + str(e))
        return formatd_xml

    # TODO check if require
    @staticmethod
    def get_xmlobj(xml_str: str) -> object:
        """ To get an xml object from an xml str """
        xmlobj = ET.fromstring(xml_str)
        return xmlobj
        # xmlobj = ET.XML(xml_str)
        # ET.indent(xmlobj)
        # print(type(ET.tostring(xmlobj, encoding='unicode')))
        # return ET.tostring(xmlobj, encoding='unicode')

    # TODO not tested
    @staticmethod
    def get_xml_fromfile(filepath: str) -> object:
        """ To get json object from an xml file """
        xmlobj = ET.parse(filepath)
        return xmlobj

    @staticmethod
    def get_xmldict(xml_str: str) -> dict:
        """ To convert xml to dict type - easy to read xml """
        # xml_asdict = xmltodict.parse(xml_str) # Use xmltodict lib from requirements
        # return xml_asdict
        pass


class JsonUtil:
    @staticmethod
    def get_json_str_from_dict(json_dict: dict, clsEncoder=None) -> str:
        """To get json str from a json dict
        """
        json_str = json.dumps(json_dict, indent=4, cls=clsEncoder)
        return json_str

    @staticmethod
    def get_json_dict_from_str(json_str: str) -> object:
        """To get json dict object from a json str
        """
        json_obj = json.loads(json_str)
        return json_obj

    @staticmethod
    def get_json_dict_from_file(filepath: str) -> object:
        """ To get json dict object from a json file
        """
        try:
            with open(filepath, mode='r') as f:
                data = json.load(f)
        except FileNotFoundError:
            assert False, 'File not found: ' + filepath
        return data

    @staticmethod
    def get_json_str_from_file(filepath: str) -> str:
        """ To get json str object from a json file
        """
        try:
            with open(filepath, mode='r') as f:
                data = json.load(f)
                data = JsonUtil.get_json_str_from_dict(data)
        except FileNotFoundError:
            assert False, 'File not found: ' + filepath
        return data

    @staticmethod
    def get_json_dict_val(json_dict, path):
        """To get value for a json path(eg: person.address.name)
        """
        keys = path.split('.')

        for key in keys:
            isCurrKeyList = True if '[' in key else False
            arrIndex = key[key.find("[") + 1:key.find("]", key.find("["))] if isCurrKeyList and key.find(
                "[") + 1 != key.find("]") else '' if isCurrKeyList else None
            calcArrIndex = 0 if arrIndex is None or arrIndex == '' else int(arrIndex)
            calcKey = key.replace(f'[{arrIndex}]', '') if isCurrKeyList else key

            if isCurrKeyList:
                json_dict = json_dict[calcKey][calcArrIndex]
            else:
                json_dict = json_dict.get(key, None)

        return json_dict

    @staticmethod
    def compare_json_val(json_dict, json_path, json_val):
        """Compare single json val and returns True/False
        """
        act_json_val = JsonUtil.get_json_dict_val(json_dict, json_path)

        isMatched = True if f"{act_json_val}" == f"{json_val}" else False
        if isMatched:
            print(f"{json_path} matched with {json_val}")
        else:
            print(f"{json_path} didnt match, actual {act_json_val}, expected {json_val}")

        return isMatched

    @staticmethod
    def compare_json_type(class_name, required_keys, json_dict=None, json_str=None):
        """class_name: Json is of which type
        required_keys: All required keys in json dict
        """
        assert json_dict is not None or json_str is not None, 'Both json params are None'

        isMatched = False
        try:
            '''Parse the json str into a dict'''
            if json_dict is None:
                json_dict = json.loads(json_str)

            '''Check if all required keys are in the dict'''
            if all(key in json_dict for key in required_keys):
                '''Try to create an object of the given class from the dict'''
                obj = class_name(**json_dict)
                isMatched = True
        except Exception as e:
            print(f"Error: {e}")
            isMatched = False

        if isMatched:
            print(f"Json type matched with {class_name} for {required_keys}")
        else:
            print(f"Json type didnt match with {class_name} for {required_keys}")

        return isMatched

    @staticmethod
    def compare_json_schema(json_dict, json_schema):
        """Compare json dict with schema
        """
        isMatched = False
        try:
            validate(json_dict, json_schema)
            isMatched = True
        except ValidationError as e:
            print(str(e))

        if isMatched:
            print(f"Json schema matched with {json_schema}")
        else:
            print(f"Json schema didnt match with {json_schema}")

        return isMatched


class FileUtil:

    @staticmethod
    def create_file(filepath: str):
        try:
            with open(filepath, mode='w') as f:
                pass
        except Exception:
            assert False, 'Exception found during file creation: ' + filepath

    @staticmethod
    def read_file(filepath: str) -> str:
        try:
            with open(filepath, mode='r') as f:
                contents = f.read()
        except FileNotFoundError:
            assert False, 'File not found: ' + filepath
        except Exception:
            assert False, 'Exception found during file reading: ' + filepath
        return contents

    @staticmethod
    def append_file(filepath: str, content):
        try:
            with open(filepath, mode='a', encoding='utf-8') as f:
                contents = f.write(content)
        except FileNotFoundError:
            assert False, 'File not found: ' + filepath
        except UnicodeEncodeError as e:
            assert False, 'UnicodeEncodeError found during file appending: ' + filepath
        except Exception:
            assert False, 'Exception found during file appending: ' + filepath

    @staticmethod
    def write_file(filepath: str, content):
        try:
            with open(filepath, mode='w') as f:
                contents = f.write(content)
        except FileNotFoundError:
            assert False, 'File not found: ' + filepath
        except UnicodeEncodeError as e:
            assert False, 'UnicodeEncodeError found during file writing: ' + filepath
        except Exception:
            assert False, 'Exception found during file writing: ' + filepath

    @staticmethod
    def replace_in_file(filepath: str, old_val: str, new_val: str):
        pass

    @staticmethod
    def archive_outputs(file_dttm):
        output_file = 'archive_{}'
        file_dttm = Commons.build_date_forfilename()
        arch_path = os.path.join(OUTPUT_DIR, output_file.format(file_dttm))

        '''Move test dir specific reports to archive dir'''
        testdirs = os.listdir(TEST_DIR)
        for td in testdirs:
            subdir = os.path.join(TEST_DIR, td)
            if os.path.isdir(subdir) and not subdir.startswith('_'):
                suboutdir = os.path.join(subdir, 'output')
                if os.path.exists(suboutdir):
                    for sd in os.listdir(suboutdir):
                        oldFile = os.path.join(suboutdir, sd)
                        filesplit = os.path.splitext(sd)
                        newFile = os.path.join(arch_path, filesplit[0] + '_' + td + filesplit[1])
                        try:
                            if not os.path.exists(arch_path):
                                os.mkdir(arch_path)
                        except Exception as e:
                            print('Exception while creating archive dir', str(e))
                        os.rename(oldFile, newFile)

        '''Move output reports to archive dir'''
        oldfiles = [f for f in os.listdir(OUTPUT_DIR) if f.lower().endswith(('.html', '.xls', '.xlsx', '.log'))]
        screenshots = runtimefiles = ''
        if os.path.exists(SCREENSHOT_DIR):
            screenshots = [f for f in os.listdir(SCREENSHOT_DIR) if f.lower().endswith('.png')]
        if os.path.exists(RUNTIME_DIR):
            runtimefiles = [f for f in os.listdir(RUNTIME_DIR) if f.lower().endswith('.xml')]
        if len(oldfiles) > 0 or len(screenshots) > 0 or len(runtimefiles) > 0:
            try:
                if not os.path.exists(arch_path):
                    os.mkdir(arch_path)
            except Exception as e:
                print('Exception while creating archive dir', str(e))

            for oldfile in oldfiles:
                try:
                    os.rename(OUTPUT_DIR + '/' + oldfile, arch_path + '/' + oldfile)
                except (PermissionError, FileNotFoundError):
                    pass
            for screenshot in screenshots:
                try:
                    os.rename(SCREENSHOT_DIR + '/' + screenshot, arch_path + '/' + screenshot)
                except (PermissionError, FileNotFoundError):
                    pass
            for runtimefile in runtimefiles:
                try:
                    os.rename(RUNTIME_DIR + '/' + runtimefile, arch_path + '/' + runtimefile)
                except (PermissionError, FileNotFoundError):
                    pass


class FTPUtil:
    ftp_server = None

    def connect_server(self, host, user, pwd):
        ftp_server = ftplib.FTP(host, user, pwd)
        ftp_server.encoding = "utf-8"

    @classmethod
    def upload_file(cls, filepath):
        with open(filepath, "rb") as file:
            # Command for Uploading the file "STOR filename"
            cls.ftp_server.storbinary(f"STOR {filepath}", file)

    @classmethod
    def download_file(cls, filepath):
        with open(filepath, "rb") as file:
            # Command for Uploading the file "STOR filename"
            cls.ftp_server.storbinary(f"STOR {filepath}", file)


class DataHandler:
    encrypt_key = "vuK4GTj-6ZVFZ4HON52Oty5Qnovh0oikWhGF500_SkQ="

    @classmethod
    def _encrypt_it(cls, decrypted_val: str) -> str:
        try:
            fernet = Fernet(cls.encrypt_key)
            encrypted_val = fernet.encrypt(decrypted_val.encode('utf-8'))
        except Exception as e:
            assert False, 'Exception during decryption ' + str(type(e))

        return encrypted_val.decode()

    @classmethod
    def decrypt_it(cls, encrypted_val: str) -> str:
        try:
            fernet = Fernet(cls.encrypt_key)
            decrypted_val = fernet.decrypt(encrypted_val).decode()
        except Exception as e:
            assert False, 'Exception during decryption ' + str(type(e))

        return decrypted_val


# var = DataHandler._encrypt_it('ABC123')
# print(var)
# var = DataHandler.decrypt_it(var)
# print(var)


class DataGeneric:
    """ Provides func to replace all defined placeholders at run time
    eg: #ABC# means user has to replace during code developement (user defined)
        {ABC} means system to replace with dynamic data during run (system dynamic)
        #{ABC} means system to replace with user defined data (variable file) """

    '''System variables, eg: {MMDDYYYY_S-1}'''
    DATE_PLACEHOLDERS = {'MMDDYYYY_S': '%m/%d/%Y'}

    # MMDDYYYY_S means '02/21/2022'

    @classmethod
    def replace_dyn(cls, data: str) -> str:
        """Provided data will be returned with all dynamic data replaced at run time
           Placeholder eg: {MMDDYYYY_S+0} """
        today = datetime.today()
        for key in cls.DATE_PLACEHOLDERS.keys():  # 'MMDDYYYY_S'
            initial_var = '{' + key  # '{MMDDYYYY_S'
            while data.count(initial_var) > 0:
                start_index = data.index(initial_var)  # 100
                end_index = data.index('}', start_index)  # 113
                actual_var = data[start_index:end_index + 1]  # '{MMDDYYYY_S-1}'
                opertr_with_val = actual_var.replace(key, '').replace('{', '').replace('}', '')  # '-1'
                only_val = opertr_with_val[1:]  # '1'

                if opertr_with_val.startswith('+'):
                    finaldate_unformtd = today + timedelta(days=float(only_val))
                else:
                    finaldate_unformtd = today + timedelta(days=-float(only_val))  # today-1

                # finaldate_formtd = final_dt_unformtd.strftime(cls.DATE_DICT[key])  # today-1 as formatted
                finaldate_formtd = Commons.format_date(finaldate_unformtd, cls.DATE_PLACEHOLDERS[key])
                data = data.replace(actual_var, str(finaldate_formtd))  # data with current placeholer replaced
        # TODO code for other placeholders
        return data

    @classmethod
    def get_vardata(cls, exlfilepath: str, sheet: str, column: str) -> dict:
        vardata = dict()
        if column is None or column == '':
            assert False, 'Column in variable file is not correct'
        else:
            worksheet = ExcelUtil.read_xlsheet(exlfilepath, sheet)
            maxcol = worksheet.max_column
            maxrow = worksheet.max_row
            col_index = None
            for c in range(2, maxcol + 1):
                if worksheet.cell(1, c).value == column:
                    col_index = c
                    break
            if col_index is None:
                assert False, 'Column in varibale file not found'
            else:
                for r in range(2, maxrow + 1):
                    vardata[worksheet.cell(r, 1).value] = worksheet.cell(r, col_index).value
        return vardata

    @classmethod
    def _replace_data_with_varfile(cls, data: str, exlfilepath: str, sheet: str, column: str) -> (str, dict):
        """Provided data will be returned with all user variables replaced at run time
           Placeholder eg: #{ITEM_HEIGHT} """
        varfiledict = DataGeneric.get_vardata(exlfilepath, sheet, column)
        for k, v in varfiledict.items():
            actual_var = '#{' + str(k) + '}'
            while data.count(actual_var) > 0:
                data = data.replace(actual_var, str(v))
        return data, varfiledict

    @classmethod
    def _replace_file_with_varfile(cls, filepath: str, exlfilepath: str, sheet: str, column: str) -> (str, dict):
        filedata = FileUtil.read_file(filepath)
        filedata, varfiledict = cls._replace_data_with_varfile(filedata, exlfilepath, sheet, column)
        FileUtil.write_file(filepath, filedata)
        return filedata, varfiledict

    @classmethod
    def replace_from_varfile(cls, exlfilepath: str, sheet: str, column: str, data: str = None, filepath: str = None) \
            -> (str, dict):
        """Replaces: data/filepath data (str) with exlfilepath data (excel).
            Returns: Replaced data (str), Var file data (dict)"""
        if data is not None:
            return cls._replace_data_with_varfile(data, exlfilepath, sheet, column)
        elif filepath is not None:
            return cls._replace_file_with_varfile(filepath, exlfilepath, sheet, column)


# data = """<ExternalSystemPurchaseOrderNbr>{MMDDYYYY_S-1} 00:01</ExternalSystemPurchaseOrderNbr>
#             <OriginFacilityAliasId>{MMDDYYYY_S+1} 123</OriginFacilityAliasId>
#             <OriginFacilityAliasId>#MMDDYYYY_S+1# 123</OriginFacilityAliasId>
#             <PickupStartDttm>{MMDDYYYY_S+0} 00:01</PickupStartDttm>
#             <PickupEndDttm>02/22/2022 00:01</PickupEndDttm>
#             <DestinationFacilityName>{MMDDYYYY_S+1} ABC</DestinationFacilityName>
#             <DestinationFacilityName>#{QTY_UOM} ABC</DestinationFacilityName>"""
#
# filepath = 'D:/Practice/AutomationService/resources/data/data_variable.xlsx'
# dataout = DataGeneric.replace_var(data, filepath, 'DO', 'DATA2')
# # print(dataout)
#
# dataout = DataGeneric.replace_dyn(dataout)
# print(dataout)
#
# filepath = 'D:/4_GPC/AutomationPrep/MASTER/AutomationService/resources/data/data_variable.xlsx'
# vardata = DataGeneric.get_var(filepath, 'DO', 'DATA1')
# print(vardata)
# print(vardata['QTY_UOM'])
# print(vardata['BU_UNIT'])

class TCPUtil:

    def __init__(self, host, port):
        self.host = host
        self.port = port

    def send_message(self, data):
        clientSocket = None
        dataFromServer = None
        try:
            clientSocket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            clientSocket.connect((self.host, self.port))
            clientSocket.send(data.encode())
            dataFromServer = clientSocket.recv(1024)
        except Exception as e:
            assert False, 'Exception while sending TCP msg: ' + str(e)
        finally:
            try:
                if clientSocket is not None:
                    clientSocket.close()
            except Exception:
                pass
        return dataFromServer
